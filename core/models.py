from django.db import models
from openpyxl.workbook import Workbook


class Author(models.Model):
    surname = models.CharField('Фамилия', max_length=255)
    first_name = models.CharField('Имя', max_length=255)
    patronymic = models.CharField('Отчество', max_length=255)
    birthday = models.DateField('Дата рождения')

    class Meta:
        verbose_name = 'Автор'
        verbose_name_plural = 'Авторы'
        ordering = ('id',)

    def __str__(self) -> str:
        return f'{self.surname} {self.first_name} {self.patronymic}'

    @staticmethod
    def to_xlsx() -> Workbook:
        report = Workbook()
        ws = report.active
        ws['A1'] = 'Фамилия'
        ws['B1'] = 'Имя'
        ws['C1'] = 'Отчество'
        ws['D1'] = 'Дата рождения'

        authors = Author.objects.all()

        for index, obj in enumerate(authors, start=2):
            ws[f'A{index}'] = obj.surname
            ws[f'B{index}'] = obj.first_name
            ws[f'C{index}'] = obj.patronymic
            ws[f'D{index}'] = obj.birthday

        return report


class Genre(models.Model):
    name = models.CharField('Название', max_length=255)

    class Meta:
        verbose_name = 'Жанр'
        verbose_name_plural = 'Жанры'
        ordering = ('id',)

    def __str__(self) -> str:
        return self.name

    @staticmethod
    def to_xlsx() -> Workbook:
        report = Workbook()
        ws = report.active
        ws['A1'] = 'Название'

        authors = Genre.objects.all()

        for index, obj in enumerate(authors, start=2):
            ws[f'A{index}'] = obj.name

        return report


class Publisher(models.Model):
    name = models.CharField('Название', max_length=255)

    class Meta:
        verbose_name = 'Издатель'
        verbose_name_plural = 'Издатели'
        ordering = ('id',)

    def __str__(self) -> str:
        return self.name

    @staticmethod
    def to_xlsx() -> Workbook:
        report = Workbook()
        ws = report.active
        ws['A1'] = 'Название'

        authors = Publisher.objects.all()

        for index, obj in enumerate(authors, start=2):
            ws[f'A{index}'] = obj.name

        return report


class Book(models.Model):
    title = models.CharField('Название', max_length=255)
    release_date = models.DateField('Дата выхода', null=True, blank=True)

    author = models.ForeignKey(Author, verbose_name='Автор', related_name='books', on_delete=models.PROTECT)
    genres = models.ManyToManyField(Genre, verbose_name='Жанры', related_name='books', blank=True)
    publisher = models.ForeignKey(
        Publisher, verbose_name='Издатель', related_name='books', blank=True, null=True, on_delete=models.SET_NULL
    )

    class Meta:
        verbose_name = 'Книга'
        verbose_name_plural = 'Книги'
        ordering = ('id',)

    def __str__(self) -> str:
        return f'{self.author} - {self.title} ({self.release_date})'

    @staticmethod
    def to_xlsx() -> Workbook:
        report = Workbook()
        ws = report.active
        ws['A1'] = 'Автор'
        ws['B1'] = 'Название'
        ws['C1'] = 'Дата выхода'
        ws['D1'] = 'Жанры'
        ws['E1'] = 'Издатель'
        authors = Book.objects.select_related('publisher', 'author').prefetch_related('genres')

        for index, obj in enumerate(authors, start=2):
            ws[f'A{index}'] = obj.title
            ws[f'B{index}'] = obj.release_date
            ws[f'C{index}'] = str(obj.author)
            ws[f'D{index}'] = ', '.join([str(genre) for genre in obj.genres.all()])
            ws[f'E{index}'] = str(obj.publisher)

        return report
